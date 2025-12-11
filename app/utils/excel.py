from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from collections import defaultdict


def export_contacts_to_excel(contacts):
    """
    Export contacts to an Excel file
    
    Args:
        contacts: List of Contact objects
        
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "通讯录"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Collect all contact method types and find max count per type
    type_max_count = defaultdict(int)
    for contact in contacts:
        type_count = defaultdict(int)
        for method in contact.methods:
            type_count[method.type] += 1
        for method_type, count in type_count.items():
            type_max_count[method_type] = max(type_max_count[method_type], count)
    
    # Define column order for contact methods
    method_type_order = ['phone', 'email', 'address', 'social']
    method_type_names = {
        'phone': '电话',
        'email': '邮箱',
        'address': '地址',
        'social': '社交媒体'
    }
    
    # Create headers
    headers = ['姓名', '是否收藏']
    method_columns = []  # Track (type, index) for each column
    
    for method_type in method_type_order:
        count = type_max_count.get(method_type, 1)
        if count == 0:
            count = 1  # At least one column per type
        for i in range(count):
            suffix = f"{i + 1}" if count > 1 else ""
            headers.append(f"{method_type_names.get(method_type, method_type)}{suffix}")
            method_columns.append((method_type, i))
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row, contact in enumerate(contacts, 2):
        # Name
        ws.cell(row=row, column=1, value=contact.name).border = thin_border
        
        # Is favorite
        ws.cell(row=row, column=2, value="是" if contact.is_favorite else "否").border = thin_border
        
        # Group methods by type
        methods_by_type = defaultdict(list)
        for method in contact.methods:
            methods_by_type[method.type].append(method.value)
        
        # Write methods
        for col, (method_type, index) in enumerate(method_columns, 3):
            values = methods_by_type.get(method_type, [])
            value = values[index] if index < len(values) else ""
            ws.cell(row=row, column=col, value=value).border = thin_border
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_contacts_from_excel(file_stream):
    """
    Import contacts from an Excel file
    
    Args:
        file_stream: File stream of the Excel file
        
    Returns:
        List of dictionaries containing contact data
    """
    wb = load_workbook(file_stream)
    ws = wb.active
    
    contacts = []
    headers = []
    
    # Read headers from first row
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        headers.append(header if header else "")
    
    # Map headers to types
    type_mapping = {
        '电话': 'phone',
        '邮箱': 'email',
        '地址': 'address',
        '社交媒体': 'social'
    }
    
    # Parse each row
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(cell_value if cell_value else "")
        
        # Skip empty rows
        if not row_data[0]:
            continue
        
        contact = {
            'name': str(row_data[0]).strip(),
            'is_favorite': str(row_data[1]).strip() == "是" if len(row_data) > 1 else False,
            'methods': []
        }
        
        # Parse contact methods
        for col, header in enumerate(headers[2:], 2):
            if col < len(row_data) and row_data[col]:
                # Determine type from header
                method_type = None
                for cn_name, en_type in type_mapping.items():
                    if header and cn_name in header:
                        method_type = en_type
                        break
                
                if method_type:
                    contact['methods'].append({
                        'type': method_type,
                        'value': str(row_data[col]).strip()
                    })
        
        contacts.append(contact)
    
    return contacts
